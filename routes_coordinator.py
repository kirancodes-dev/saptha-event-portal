import csv, datetime, io, random, secrets, string
from io import StringIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import (Blueprint, Response, flash, jsonify,
                   redirect, render_template, request, session)
from google.cloud import firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from werkzeug.security import generate_password_hash
from models import db
from utils import login_required, role_required, log_action, safe_int
from utils_email import (send_appointment_email, send_broadcast_email,
                         send_credentials_email, send_ticket_email, send_result_email)
try:
    from utils_whatsapp import (send_room_assignment_whatsapp, send_result_whatsapp,
        send_elimination_whatsapp, send_broadcast_whatsapp,
        send_staff_credentials_whatsapp, send_ticket_whatsapp)
    _WA = True
except ImportError:
    _WA = False

coord_bp    = Blueprint('coordinator', __name__, url_prefix='/coordinator')
COORD_ROLES = ['ClubSPOC', 'Coordinator', 'SuperAdmin', 'Super Admin']

def _wa(fn, *a, **kw):
    if _WA:
        try: return fn(*a, **kw)
        except Exception: pass
    return False

def _phone(reg):
    return (reg.get('lead_phone') or reg.get('phone') or
            (reg.get('members') or [{}])[0].get('phone', ''))

def _ff(f, op, v):
    return FieldFilter(f, op, v)


# 1. DASHBOARD
@coord_bp.route('/dashboard')
@login_required
@role_required(COORD_ROLES)
def dashboard():
    user_role = session.get('role', '')
    user_email = session.get('user_id')
    club_category = session.get('category', 'General')
    try:
        is_super = user_role in ('SuperAdmin', 'Super Admin') or club_category == 'All'
        if is_super:
            events_ref = (db.collection('events')
                           .order_by('created_at', direction=firestore.Query.DESCENDING)
                           .stream())
        else:
            events_ref = (db.collection('events')
                           .where(filter=_ff('created_by_email', '==', user_email))
                           .stream())
        events = []; total_regs = 0; total_staff = 0
        for e in events_ref:
            d = e.to_dict(); d['id'] = e.id
            total_regs  += d.get('registration_count', 0)
            total_staff += len(d.get('staff', []))
            regs = (db.collection('registrations')
                      .where(filter=_ff('event_id', '==', e.id)).stream())
            d['scored_teams'] = sum(
                1 for r in regs
                if not r.to_dict().get('is_eliminated') and r.to_dict().get('scores'))
            events.append(d)
    except Exception as exc:
        flash(f"Dashboard error: {exc}", "danger")
        events = []; total_regs = 0; total_staff = 0
    return render_template('coordinator/dashboard.html',
        events=events, club_category=club_category,
        total_regs=total_regs, total_staff=total_staff,
        user_name=session.get('name'))


# 2. VIEW REGISTRATIONS (all students for an event)
@coord_bp.route('/registrations/<event_id>')
@login_required
@role_required(COORD_ROLES)
def view_registrations(event_id):
    try:
        event_doc = db.collection('events').document(event_id).get()
        if not event_doc.exists:
            flash("Event not found.", "danger")
            return redirect('/coordinator/dashboard')
        event = event_doc.to_dict(); event['id'] = event_id
        regs_raw = (db.collection('registrations')
                      .where(filter=_ff('event_id', '==', event_id)).stream())
        registrations = sorted(
            [dict(r.to_dict(), id=r.id) for r in regs_raw],
            key=lambda x: x.get('registered_at', ''), reverse=True)
        return render_template('coordinator/registrations.html',
            event=event, registrations=registrations, total=len(registrations))
    except Exception as exc:
        flash(f"Error loading registrations: {exc}", "danger")
        return redirect('/coordinator/dashboard')


# 3. CREATE EVENT
@coord_bp.route('/create_event', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def create_event():
    try:
        overview = request.form.get('overview', '').strip()
        criteria_list = [c.strip() for c in
            request.form.get('criteria', 'Overall Score').split(',') if c.strip()] or ['Overall Score']
        media_urls = [u.strip() for u in request.form.getlist('media_urls[]') if u.strip()]
        category = (session.get('category') if session.get('category') not in (None, 'All')
                    else request.form.get('category', 'General'))
        db.collection('events').add({
            'title':              request.form.get('title', '').strip(),
            'date':               request.form.get('date'),
            'deadline':           request.form.get('deadline'),
            'venue':              request.form.get('venue', '').strip(),
            'description':        overview[:120] + '...' if len(overview) > 120 else overview,
            'overview':           overview,
            'rules':              request.form.get('rules', ''),
            'prizes':             request.form.get('prizes', ''),
            'category':           category,
            'media_urls':         media_urls,
            'banner_url':         media_urls[0] if media_urls else '',
            'entry_fee':          safe_int(request.form.get('entry_fee', 0)),
            'is_team_event':      request.form.get('is_team') == 'on',
            'judging_criteria':   criteria_list,
            'status':             'active',
            'active_round':       1,
            'registration_count': 0,
            'staff':              [],
            'created_by':         session.get('name'),
            'created_by_email':   session.get('user_id'),
            'created_at':         datetime.datetime.utcnow(),
        })
        log_action(db, "EVENT_CREATED",
                   f"{session.get('user_id')} created '{request.form.get('title')}'")
        flash("Event created successfully!", "success")
    except Exception as exc:
        flash(f"Error creating event: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 4. EDIT EVENT
@coord_bp.route('/edit_event/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def edit_event(event_id):
    try:
        overview = request.form.get('overview', '').strip()
        criteria_list = [c.strip() for c in
            request.form.get('criteria', 'Overall Score').split(',') if c.strip()] or ['Overall Score']
        media_urls = [u.strip() for u in request.form.getlist('media_urls[]') if u.strip()]
        db.collection('events').document(event_id).update({
            'title':            request.form.get('title', '').strip(),
            'date':             request.form.get('date'),
            'deadline':         request.form.get('deadline'),
            'venue':            request.form.get('venue', '').strip(),
            'overview':         overview,
            'description':      overview[:120] + '...' if len(overview) > 120 else overview,
            'rules':            request.form.get('rules', ''),
            'prizes':           request.form.get('prizes', ''),
            'judging_criteria': criteria_list,
            'media_urls':       media_urls,
            'banner_url':       media_urls[0] if media_urls else '',
            'entry_fee':        safe_int(request.form.get('entry_fee', 0)),
            'is_team_event':    request.form.get('is_team') == 'on',
        })
        log_action(db, "EVENT_EDITED", f"Event {event_id} by {session.get('user_id')}")
        flash("Event updated!", "success")
    except Exception as exc:
        flash(f"Update error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 5. DELETE EVENT
@coord_bp.route('/delete_event/<event_id>')
@login_required
@role_required(COORD_ROLES)
def delete_event(event_id):
    try:
        db.collection('events').document(event_id).delete()
        for r in (db.collection('registrations')
                    .where(filter=_ff('event_id', '==', event_id)).stream()):
            r.reference.delete()
        for s in (db.collection('form_submissions')
                    .where(filter=_ff('event_id', '==', event_id)).stream()):
            s.reference.delete()
        db.collection('event_forms').document(event_id).delete()
        log_action(db, "EVENT_DELETED", f"Event {event_id} by {session.get('user_id')}")
        flash("Event deleted.", "warning")
    except Exception as exc:
        flash(f"Delete error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 6. ASSIGN STAFF
@coord_bp.route('/assign_staff/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def assign_staff(event_id):
    try:
        name  = request.form.get('name', '').strip()
        email = request.form.get('email', '').lower().strip()
        role  = request.form.get('role', '')
        phone = request.form.get('phone', '').strip()
        if not name or not email or not role:
            flash("Name, email and role are required.", "warning")
            return redirect('/coordinator/dashboard')
        event_doc   = db.collection('events').document(event_id).get()
        event_title = event_doc.to_dict().get('title', 'Event') if event_doc.exists else 'Event'
        user_ref    = db.collection('users').document(email)
        if not user_ref.get().exists:
            alphabet = string.ascii_letters + string.digits
            raw_pw   = ''.join(secrets.choice(alphabet) for _ in range(10))
            user_ref.set({'email': email, 'name': name, 'role': role, 'phone': phone,
                'category': session.get('category', 'General'),
                'password': generate_password_hash(raw_pw),
                'created_at': datetime.datetime.now().strftime("%Y-%m-%d"),
                'needs_password_reset': True})
            send_credentials_email(email, name, role, raw_pw, session.get('category', 'General'))
            _wa(send_staff_credentials_whatsapp, phone, name, role, event_title, email, raw_pw)
            flash(f"Account created for {name}. Credentials emailed.", "success")
        else:
            if user_ref.get().to_dict().get('role') == 'Student':
                user_ref.update({'role': role})
            send_appointment_email(email, name, role, event_title)
            flash(f"{name} already had an account. Role updated.", "info")
        db.collection('events').document(event_id).update({
            'staff': firestore.ArrayUnion([{'name': name, 'email': email, 'role': role}])})
        log_action(db, "STAFF_ASSIGNED", f"{name} ({role}) to event {event_id}")
        flash(f"{name} appointed as {role}!", "success")
    except Exception as exc:
        flash(f"Assignment error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 7. ALLOCATE ROOMS
@coord_bp.route('/allocate_rooms/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def allocate_rooms(event_id):
    try:
        room_names = request.form.getlist('room_name[]')
        capacities = [safe_int(c) for c in request.form.getlist('capacity[]')]
        event_data = db.collection('events').document(event_id).get().to_dict()
        judges     = [s for s in event_data.get('staff', []) if s.get('role') == 'Judge']
        if len(judges) < len(room_names):
            flash(f"Need {len(room_names)} judges but only {len(judges)} assigned.", "warning")
            return redirect('/coordinator/dashboard')
        random.shuffle(judges)
        regs = [r for r in (db.collection('registrations')
                              .where(filter=_ff('event_id', '==', event_id)).stream())
                if not r.to_dict().get('is_eliminated')]
        random.shuffle(regs)
        reg_index = 0
        for i, (room, cap) in enumerate(zip(room_names, capacities)):
            judge = judges[i]
            for _ in range(cap):
                if reg_index >= len(regs): break
                db.collection('registrations').document(regs[reg_index].id).update({
                    'assigned_room': room.strip(),
                    'assigned_judge_email': judge['email'],
                    'assigned_judge_name': judge['name']})
                reg_index += 1
        log_action(db, "ROOMS_ALLOCATED",
                   f"{reg_index} teams in {len(room_names)} rooms for event {event_id}")
        flash(f"{reg_index} teams allocated across {len(room_names)} rooms!", "success")
    except Exception as exc:
        flash(f"Allocation error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 8. TRIGGER REMINDERS
@coord_bp.route('/trigger_reminders/<event_id>')
@login_required
@role_required(COORD_ROLES)
def trigger_reminders(event_id):
    try:
        event_data  = db.collection('events').document(event_id).get().to_dict()
        round_num   = event_data.get('active_round', 1)
        email_count = wa_count = 0
        for r in (db.collection('registrations')
                    .where(filter=_ff('event_id', '==', event_id)).stream()):
            d = r.to_dict()
            if d.get('is_eliminated') or not d.get('assigned_room'): continue
            lead_email = d.get('lead_email', ''); lead_name = d.get('lead_name', 'Participant')
            room = d.get('assigned_room', 'TBD'); judge = d.get('assigned_judge_name', 'TBD')
            phone = _phone(d)
            body = (f"ROUND {round_num} ALERT\n\nHello {lead_name},\n\n"
                    f"You are in Round {round_num} of '{event_data['title']}'.\n"
                    f"Room: {room}\nJudge: {judge}\n\nBest of luck!")
            send_broadcast_email([lead_email], f"Round {round_num} Details",
                                  body, event_data['title'])
            email_count += 1
            if phone and _wa(send_room_assignment_whatsapp, phone, lead_name,
                             event_data['title'], round_num, room, judge):
                wa_count += 1
        flash(f"Round {round_num} alerts — {email_count} emails, {wa_count} WhatsApp.", "success")
    except Exception as exc:
        flash(f"Reminder error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 9. PROMOTE ROUND / ELIMINATION
@coord_bp.route('/promote_round/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def promote_round(event_id):
    try:
        cutoff    = float(request.form.get('cutoff_score', 0))
        event_ref = db.collection('events').document(event_id)
        event_d   = event_ref.get().to_dict()
        cur_round = event_d.get('active_round', 1)
        nxt_round = cur_round + 1
        promoted = eliminated = 0
        for reg in (db.collection('registrations')
                      .where(filter=_ff('event_id', '==', event_id)).stream()):
            rd = reg.to_dict()
            if rd.get('is_eliminated'): continue
            scores = rd.get('scores', {})
            total  = sum(safe_int(s.get('total', 0)) for s in scores.values())
            rr     = db.collection('registrations').document(reg.id)
            if total >= cutoff:
                rr.update({'current_round': nxt_round, 'scores': firestore.DELETE_FIELD,
                            'assigned_room': None, 'assigned_judge_email': None,
                            'assigned_judge_name': None})
                promoted += 1
            else:
                rr.update({'is_eliminated': True})
                eliminated += 1
                _wa(send_elimination_whatsapp, _phone(rd), rd.get('lead_name', ''),
                    event_d.get('title', 'Event'), cur_round)
        event_ref.update({'active_round': nxt_round})
        log_action(db, "ROUND_PROMOTED",
                   f"Event {event_id}: {promoted} promoted, {eliminated} eliminated. Cutoff={cutoff}")
        flash(f"Round {nxt_round}: {promoted} advanced, {eliminated} eliminated.", "success")
    except Exception as exc:
        flash(f"Promotion error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 10. BROADCAST
@coord_bp.route('/broadcast/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def broadcast_message(event_id):
    try:
        subject   = request.form.get('subject', '').strip()
        message   = request.form.get('message', '').strip()
        event_doc = db.collection('events').document(event_id).get()
        if not event_doc.exists: return redirect('/coordinator/dashboard')
        event_title = event_doc.to_dict().get('title', 'Event')
        regs        = list(db.collection('registrations')
                           .where(filter=_ff('event_id', '==', event_id)).stream())
        email_list  = list({r.to_dict().get('lead_email') for r in regs
                            if r.to_dict().get('lead_email')})
        phone_list  = list({_phone(r.to_dict()) for r in regs if _phone(r.to_dict())})
        send_broadcast_email(email_list, subject, message, event_title)
        wa_sent = 0
        if _WA:
            try:
                res = send_broadcast_whatsapp(phone_list, event_title, subject, message)
                wa_sent = res.get('sent', 0)
            except Exception: pass
        flash(f"Sent to {len(email_list)} emails, {wa_sent} WhatsApp.", "success")
    except Exception as exc:
        flash(f"Broadcast error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 11. PUBLISH RESULTS
@coord_bp.route('/publish_results/<event_id>', methods=['POST'])
@login_required
@role_required(COORD_ROLES)
def publish_results(event_id):
    try:
        event_ref   = db.collection('events').document(event_id)
        event_data  = event_ref.get().to_dict() or {}
        event_title = event_data.get('title', 'Event')
        leaderboard = []
        for r in (db.collection('registrations')
                    .where(filter=_ff('event_id', '==', event_id)).stream()):
            d = r.to_dict()
            if d.get('is_eliminated'): continue
            scores = d.get('scores', {})
            if not scores: continue
            avg = round(sum(safe_int(s.get('total', 0)) for s in scores.values()) / len(scores), 2)
            leaderboard.append({'team_name': d.get('team_name'), 'lead_name': d.get('lead_name'),
                                 'email': d.get('lead_email'), 'phone': _phone(d), 'score': avg})
        leaderboard.sort(key=lambda x: x['score'], reverse=True)
        for idx, winner in enumerate(leaderboard[:3], start=1):
            send_result_email(winner['email'], winner['lead_name'], event_title, idx, winner['score'])
            if winner.get('phone'):
                _wa(send_result_whatsapp, winner['phone'], winner['lead_name'],
                    event_title, idx, winner['score'])
        event_ref.update({'status': 'completed', 'winners': leaderboard[:3],
                          'completed_at': datetime.datetime.utcnow()})
        log_action(db, "RESULTS_PUBLISHED", f"Event {event_id} by {session.get('user_id')}")
        flash("Results published! Top 3 notified.", "success")
    except Exception as exc:
        flash(f"Publish error: {exc}", "danger")
    return redirect('/coordinator/dashboard')


# 12. EXPORT CSV
@coord_bp.route('/export_registrations/<event_id>')
@login_required
@role_required(COORD_ROLES)
def export_registrations(event_id):
    try:
        event_doc  = db.collection('events').document(event_id).get()
        event_data = event_doc.to_dict() if event_doc.exists else {}
        is_team    = event_data.get('is_team_event', False)
        regs       = list(db.collection('registrations')
                          .where(filter=_ff('event_id', '==', event_id)).stream())
        output = StringIO()
        writer = csv.writer(output)
        header = ['Ticket ID', 'Lead Name', 'Email', 'Phone', 'USN', 'Team Name',
                  'Payment Status', 'Amount (Rs)', 'Room', 'Judge', 'Round',
                  'Status', 'Attendance', 'Check-in Time', 'Registered At']
        if is_team:
            for i in range(2, 6):
                header += [f'M{i} Name', f'M{i} Email', f'M{i} Phone', f'M{i} USN']
        writer.writerow(header)
        for r in regs:
            d      = r.to_dict()
            status = 'Eliminated' if d.get('is_eliminated') else 'Active'
            row    = [d.get('reg_id', r.id), d.get('lead_name', ''), d.get('lead_email', ''),
                      d.get('lead_phone', ''), d.get('lead_usn', ''),
                      d.get('team_name', 'Individual'), d.get('payment_status', 'Free'),
                      d.get('amount_paid', 0), d.get('assigned_room', ''),
                      d.get('assigned_judge_name', ''), d.get('current_round', 1),
                      status, d.get('attendance', 'Pending'),
                      d.get('checkin_time', ''), d.get('registered_at', '')]
            if is_team:
                mems = [m for m in d.get('members', [])
                        if m.get('email', '') != d.get('lead_email', '')]
                for i in range(4):
                    m = mems[i] if i < len(mems) else {}
                    row += [m.get('name', ''), m.get('email', ''),
                             m.get('phone', ''), m.get('usn', '')]
            writer.writerow(row)
        title = event_data.get('title', 'Event').replace(' ', '_')
        return Response(output.getvalue(), mimetype='text/csv',
            headers={"Content-Disposition": f"attachment; filename={title}_Registrations.csv"})
    except Exception as exc:
        flash(f"CSV export error: {exc}", "danger")
        return redirect('/coordinator/dashboard')


# 13. EXPORT EXCEL (branded, all team members expanded)
@coord_bp.route('/export_excel/<event_id>')
@login_required
@role_required(COORD_ROLES)
def export_excel(event_id):
    try:
        event_doc  = db.collection('events').document(event_id).get()
        event_data = event_doc.to_dict() if event_doc.exists else {}
        is_team    = event_data.get('is_team_event', False)
        regs       = list(db.collection('registrations')
                          .where(filter=_ff('event_id', '==', event_id)).stream())
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Registrations"
        hf    = PatternFill("solid", fgColor="0D2D62")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        haln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dfont = Font(size=10)
        daln  = Alignment(vertical="center")
        afill = PatternFill("solid", fgColor="EEF3FA")
        thin  = Side(style="thin", color="CCCCCC")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        event_title = event_data.get('title', 'Event')
        last_col    = 15 + (16 if is_team else 0)
        last_ltr    = get_column_letter(last_col)
        # Row 1 — title banner
        ws.merge_cells(f'A1:{last_ltr}1')
        c = ws['A1']
        c.value = f"Sapthagiri NPS University  |  {event_title}  |  Registrations"
        c.font  = Font(bold=True, color="F37021", size=13)
        c.fill  = PatternFill("solid", fgColor="0D2D62")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        # Row 2 — subtitle
        ws.merge_cells(f'A2:{last_ltr}2')
        c2 = ws['A2']
        c2.value = (f"Date: {event_data.get('date','TBD')}  |  "
                    f"Venue: {event_data.get('venue','SNPSU')}  |  "
                    f"Total Registrations: {len(regs)}  |  "
                    f"Exported: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}")
        c2.font  = Font(italic=True, color="FFFFFF", size=9)
        c2.fill  = PatternFill("solid", fgColor="1A3A6B")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16
        # Row 3 — headers
        headers = ['S.No', 'Ticket ID', 'Lead Name', 'Email', 'Phone', 'USN',
                   'Team Name', 'Payment', 'Amount (Rs)', 'Room', 'Round',
                   'Status', 'Attendance', 'Check-in', 'Registered At']
        if is_team:
            for i in range(2, 6):
                headers += [f'M{i} Name', f'M{i} Email', f'M{i} Phone', f'M{i} USN']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = hfont; cell.fill = hf; cell.alignment = haln; cell.border = bdr
        ws.row_dimensions[3].height = 26
        # Data rows
        for ri, r in enumerate(regs, 1):
            d      = r.to_dict()
            status = 'Eliminated' if d.get('is_eliminated') else 'Active'
            fill   = afill if ri % 2 == 0 else PatternFill()
            row    = [ri, d.get('reg_id', r.id), d.get('lead_name', ''),
                      d.get('lead_email', ''), d.get('lead_phone', ''),
                      d.get('lead_usn', ''), d.get('team_name', 'Individual'),
                      d.get('payment_status', 'Free'), d.get('amount_paid', 0),
                      d.get('assigned_room', ''), d.get('current_round', 1),
                      status, d.get('attendance', 'Pending'),
                      d.get('checkin_time', ''), d.get('registered_at', '')]
            if is_team:
                mems = [m for m in d.get('members', [])
                        if m.get('email', '') != d.get('lead_email', '')]
                for i in range(4):
                    m = mems[i] if i < len(mems) else {}
                    row += [m.get('name', ''), m.get('email', ''),
                             m.get('phone', ''), m.get('usn', '')]
            xlsx_row = ri + 3
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=xlsx_row, column=ci, value=val)
                cell.font = dfont; cell.alignment = daln; cell.border = bdr
                if fill.fill_type: cell.fill = fill
            ws.row_dimensions[xlsx_row].height = 17
        widths = [5, 22, 20, 28, 14, 15, 20, 12, 10, 12, 7, 12, 12, 10, 20]
        if is_team: widths += [18, 26, 13, 13] * 4
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A4'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        clean = event_title.replace(' ', '_')
        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={clean}_Registrations.xlsx"})
    except Exception as exc:
        flash(f"Excel export error: {exc}", "danger")
        return redirect('/coordinator/dashboard')


# 14. WALK-IN / ON-SPOT
@coord_bp.route('/on_spot')
@login_required
@role_required(['EventCoordinator', 'SuperAdmin', 'Super Admin'])
def on_spot_form():
    events = [{'id': e.id, 'title': e.to_dict().get('title')}
              for e in (db.collection('events')
                          .where(filter=_ff('status', '==', 'active')).stream())]
    return render_template('coordinator/on_spot.html', events=events)


@coord_bp.route('/process_walkin', methods=['POST'])
@login_required
@role_required(['EventCoordinator', 'SuperAdmin', 'Super Admin'])
def process_walkin():
    try:
        import time as _time
        event_id = request.form.get('event_id', '')
        email    = request.form.get('email', '').lower().strip()
        name     = request.form.get('name', '').strip()
        usn      = request.form.get('usn', '').upper().strip()
        phone    = request.form.get('phone', '').strip()
        if not event_id or not email or not name:
            flash("Event, name and email are required.", "warning")
            return redirect('/coordinator/on_spot')
        user_ref = db.collection('users').document(email)
        if not user_ref.get().exists:
            user_ref.set({'email': email, 'name': name, 'phone': phone, 'usn': usn,
                'role': 'Student', 'password': generate_password_hash('Welcome@123'),
                'created_at': datetime.datetime.now().strftime("%Y-%m-%d"),
                'needs_password_reset': True})
        reg_id      = f"REG-{int(_time.time() * 1000)}"
        event_doc   = db.collection('events').document(event_id).get().to_dict() or {}
        event_title = event_doc.get('title', 'Event')
        db.collection('registrations').document(reg_id).set({
            'reg_id': reg_id, 'event_id': event_id, 'event_title': event_title,
            'lead_name': name, 'lead_email': email, 'lead_usn': usn, 'lead_phone': phone,
            'team_name': 'Walk-in', 'members': [],
            'status': 'Confirmed', 'payment_status': 'Paid',
            'amount_paid': event_doc.get('entry_fee', 0),
            'payment_mode': request.form.get('payment_mode', 'Cash'),
            'registered_at': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'attendance': 'Present',
            'checkin_time': datetime.datetime.now().strftime("%H:%M:%S"),
            'is_eliminated': False, 'current_round': event_doc.get('active_round', 1)})
        send_ticket_email(email, name, event_title, reg_id)
        _wa(send_ticket_whatsapp, phone, name, event_title, reg_id,
            event_doc.get('date', ''), event_doc.get('venue', ''))
        log_action(db, "WALKIN_REGISTERED", f"Walk-in {email} for event {event_id}")
        flash(f"Walk-in for {name} registered. Ticket sent.", "success")
    except Exception as exc:
        flash(f"Walk-in error: {exc}", "danger")
    return redirect('/coordinator/on_spot')


# 15. QR SCANNER
@coord_bp.route('/scanner')
@login_required
@role_required(['EventCoordinator', 'SuperAdmin', 'Super Admin'])
def scanner_selector():
    events = [{'id': e.id, 'title': e.to_dict().get('title')}
              for e in (db.collection('events')
                          .where(filter=_ff('status', '==', 'active')).stream())]
    return render_template('coordinator/scanner_selector.html', events=events)


@coord_bp.route('/scan/<event_id>')
@login_required
@role_required(['EventCoordinator', 'SuperAdmin', 'Super Admin'])
def scan_page(event_id):
    doc = db.collection('events').document(event_id).get()
    if not doc.exists: return redirect('/coordinator/scanner')
    return render_template('coordinator/scan.html',
                            event_id=event_id,
                            event_title=doc.to_dict().get('title'))


@coord_bp.route('/get_ticket/<reg_id>')
@login_required
def get_ticket(reg_id):
    reg = db.collection('registrations').document(reg_id).get()
    if not reg.exists:
        return jsonify({'status': 'error', 'message': 'INVALID TICKET'})
    return jsonify({'status': 'success', 'data': reg.to_dict()})


@coord_bp.route('/mark_attendance_granular', methods=['POST'])
@login_required
def mark_attendance_granular():
    try:
        data         = request.json or {}
        reg_id       = data.get('reg_id')
        present_usns = data.get('present_usns', [])
        if not reg_id:
            return jsonify({'status': 'error', 'message': 'Missing reg_id'})
        reg_ref  = db.collection('registrations').document(reg_id)
        reg_data = reg_ref.get().to_dict()
        if not reg_data:
            return jsonify({'status': 'error', 'message': 'Registration not found'})
        if reg_data.get('payment_status') == 'Pending':
            return jsonify({'status': 'error', 'message': 'PAYMENT PENDING!'})
        members = reg_data.get('members', [])
        for m in members:
            m['attendance'] = 'Present' if m.get('usn') in present_usns else 'Absent'
        reg_ref.update({'members': members,
                        'attendance': 'Present' if present_usns else 'Absent',
                        'checkin_time': datetime.datetime.now().strftime("%H:%M:%S")})
        log_action(db, "ATTENDANCE_MARKED",
                   f"Reg {reg_id}: {len(present_usns)} members marked present")
        return jsonify({'status': 'success',
                        'message': f"{len(present_usns)} members marked present."})
    except Exception as exc:
        return jsonify({'status': 'error', 'message': str(exc)})


# 16. CERTIFICATE
@coord_bp.route('/certificate/<reg_id>/<usn>')
def generate_certificate(reg_id, usn):
    reg_doc = db.collection('registrations').document(reg_id).get()
    if not reg_doc.exists: return "Registration not found.", 404
    data = reg_doc.to_dict()
    student_name = None
    if data.get('lead_usn') == usn:
        if data.get('attendance') != 'Present':
            return "Certificate denied: marked Absent.", 403
        student_name = data.get('lead_name')
    else:
        for m in data.get('members', []):
            if m.get('usn') == usn:
                if m.get('attendance') != 'Present':
                    return "Certificate denied: marked Absent.", 403
                student_name = m.get('name')
                break
    if not student_name:
        return "Student not found in this registration.", 404
    event_data = db.collection('events').document(data['event_id']).get().to_dict()
    return render_template('participant/certificate.html',
                            student_name=student_name, event=event_data)